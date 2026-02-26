import openpyxl
import sys

output_file = r'c:\Users\Diegu\People\excel_dump.txt'

try:
    with open(output_file, 'w', encoding='utf-8') as f:
        wb = openpyxl.load_workbook(r'c:\Users\Diegu\People\public\RENDIMIENTO OPERACIONAL.xlsx', data_only=True)
        f.write(f'HOJAS ENCONTRADAS: {wb.sheetnames}\n')

        sheets_to_read = ['Petshop', 'Grooming', 'Clinica']
        
        for sheet_name in sheets_to_read:
            if sheet_name not in wb.sheetnames:
                continue
                
            f.write(f'\n{"="*40}\n')
            f.write(f'HOJA: {sheet_name}\n')
            f.write(f'{"="*40}\n')
            
            sheet = wb[sheet_name]
            
            # Leer primeras 100 filas
            max_row = min(150, sheet.max_row + 1)
            max_col = min(20, sheet.max_column + 1)
            
            for row in range(1, max_row):
                vals = []
                has_data = False
                for col in range(1, max_col):
                    cell = sheet.cell(row=row, column=col).value
                    val_str = str(cell).strip() if cell is not None else ""
                    # Truncar si es muy largo para legibilidad
                    if len(val_str) > 60:
                        val_str = val_str[:57] + "..."
                    vals.append(val_str)
                    if val_str:
                        has_data = True
                
                if has_data:
                    # Formato simple separado por pipes
                    row_str = " | ".join(vals)
                    f.write(f"R{row:02d}: {row_str}\n")

    print(f"Archivo generado correctamente: {output_file}")

except Exception as e:
    print(f"ERROR: {e}")
